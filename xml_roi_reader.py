import xml.etree.ElementTree as ET
from os import listdir
from os.path import isfile, join
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

class ROI:
    def __init__(self, id, file_name, narrative):
        self.id = id
        self.file_name = file_name
        self.narrative = narrative
        
# Directory of ROI XML files
#xml_dir_path = "/path/to/roi/xml"
#linux_separator = "/"
win_path = 'C:\\work\\roi\\xml'
win_separator = '\\'

def generate_roi_excel(rois, RESULTS_DIR):
    print(f"Generating XLSX")
    #logger = logging.getLogger()

    try:
        # Create results data
        rows = []
        for roi in rois:
            row_data = []
            row_data.append(roi.id)
            row_data.append(roi.narrative)
            rows.append(row_data)

        # Create results XLSX
        wb = Workbook()
        #job_id = datetime.now().strftime('%m%d%Y_%H%M%S')
        xlsx_out = RESULTS_DIR + "\\rois_sample1.xlsx"
        ws1 = wb.active
        ws1.title = "Results"

        # Create column headers
        result_headers = ['ROI_ID', 'Narrative']
        #result_headers.extend(headers)
        ws1.append(result_headers)

        ws1['A1'].font = Font(bold=True)
        ws1['B1'].font = Font(bold=True)

        for row in rows:
            ws1.append(row)

        # Save XLSX
        wb.save(filename = xlsx_out)

        print(f"Wrote Excel results to: {xlsx_out}")
    except Exception as e:
        print(f"Exception {e}")
        return None, str(e)

# List of ROI objects
def extract_rois(win_path):
    rois = []
    for f in listdir(win_path):
        if isfile(join(win_path, f)):
            if re.match(r'^batch\-\d{3}\.[x]', f):
                tree = ET.parse(win_path + win_separator + f)
                root = tree.getroot()
                for i, element in enumerate(root.findall('DOC')):
                    roi_id = element.find('ID').text # ROI number
                    narrative = element.find('TXT').text  # Narratives
                    roi = ROI(roi_id, f, narrative)
                    print(f'-------------------\nROI: {roi_id}\n' + narrative + "\n")
                    rois.append(roi)
    return rois
             
def extract_rois_to_xlsx(win_path):
    rois = []
    for f in listdir(win_path):
        if isfile(join(win_path, f)):
            if re.match(r'^batch\-\d{3}\.[x]', f):
                tree = ET.parse(win_path + win_separator + f)
                root = tree.getroot()
                for i, element in enumerate(root.findall('DOC')):
                    roi_id = element.find('ID').text # ROI number
                    narrative = element.find('TXT').text  # Narratives
                    roi = ROI(roi_id, f, narrative)
                    print(f'-------------------\nROI: {roi_id}\n' + narrative + "\n")
                    rois.append(roi)
                    
    generate_roi_excel(rois, 'C:\\work\\roi\\xml')

   

# Run
print(f"doing it")
extract_rois_to_xlsx("C:\\work\\roi\\xml")
