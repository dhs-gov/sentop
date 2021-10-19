import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from . import sentop_log
from . import globalutils
from .DataIn import DataIn  # Note that '.' is used to denote a Python class (not module).
from .DataType import DataType
from . import text_validator

class XlsxData (DataIn):

    """ The incoming XLSX data.

    Attributes:
        kms_id                  The KMS ID (not used for file URLs).
        sentop_id               The SENTOP ID.
        row_id_list             The list of row IDs.
        data_list               The corpus (i.e., list of documents) to be analyzed.
        all_stop_words          The list of all stop words including user-defined stop words.
        annotation              The user-defined annotation.
        xlsx_data_table         The spreadsheet data minus headers (XLSX only).
        xlsx_headers_row        The row number of the lowest-level headers (XLSX only).
    """
        
    def __init__(self, kms_id, sentop_id, row_id_list, data_list, all_stop_words, annotation, table_data, table_col_headers, table_headers_row_index):

        DataIn.__init__(self, DataType.XLSX, kms_id, sentop_id, row_id_list, data_list, all_stop_words, annotation) 
        self.table_data = table_data
        self.table_col_headers = table_col_headers
        self.table_headers_row_index = table_headers_row_index

    def show_info(self):
        sentlog = sentop_log.SentopLog()
        sentlog.info_h2(f"Data Summary")
        if self.data_type:
            sentlog.info_keyval(f"Data Type|{self.data_type}")
        else:
            sentlog.error("No data type for this object.")

        if self.kms_id:
            sentlog.info_keyval(f"KMS ID|{self.kms_id}")
        else:
            sentlog.error("No KMS ID for this data.")

        if self.sentop_id:
            sentlog.info_keyval(f"SENTOP ID|{self.sentop_id}")
        else:
            sentlog.error("No SENTOP ID for this data.")

        #if self.row_id_list:
        #    sentlog.info_keyval(f"Num Valid Data IDs|{len(self.row_id_list)}")
        #else:
        #   sentlog.error("No row ID list for this data.")

        if self.data_list:
            sentlog.info_keyval(f"Num Valid Data|{len(self.data_list)}")
        else:
            sentlog.error("No data list for this data.")

        if self.all_stop_words:
            sentlog.info_keyval(f"All stop words|{len(self.all_stop_words)}")
        else:
            sentlog.error("No stop words for this data.")

        if self.annotation:
            sentlog.info_keyval(f"Annotation|{self.annotation}")
        else:
            sentlog.error("No annotation for this data.")

        if self.table_data:
            sentlog.info_keyval(f"len(table_data)|{len(self.table_data)}")
        else:
            sentlog.error("No data table for this data.")

        if self.table_col_headers:
            sentlog.info_keyval(f"len(table_col_headers)|{len(self.table_col_headers)}")
        else:
            sentlog.error("No table col headers for this data.")
            
        if self.table_headers_row_index:
            sentlog.info_keyval(f"table_headers_row_index|{self.table_headers_row_index}")
        else:
            sentlog.error("No header rows for this data.")

        sentlog.info_p(f"")


def get_user_stopwords(config_ws, row, column):
    sentlog = sentop_log.SentopLog()
    user_stop_words = []
    try:
        done = False
        while not done:
            stop_word = config_ws.cell(row, column).value
            if stop_word:
                sentlog.info_keyval(f"Found stop word|{stop_word}")
                user_stop_words.append(stop_word)
                row += 1 
            else:
                done = True

        return user_stop_words

    except Exception as e:
        globalutils.show_stack_trace(str(e))
        sentlog.error("Error getting user user stop words.")
        return None


def get_data(req, bytes, sentop_id, kms_id):

    sentlog = sentop_log.SentopLog()
    try:

        # ------------------------ GET XLSX WORKBOOK -------------------------

        xlsx = io.BytesIO(bytes)
        wb = load_workbook(xlsx)

        # ------------------------ GET SENTOP CONFIG SHEET -------------------------
        # XLSX files MUST have a SENTOP Config sheet.

        config_ws = None
        try:
            config_ws = wb.get_sheet_by_name('SENTOP Config')
            if not config_ws:
                sentlog.error(f"SENTOP Config sheet not found. Aborting.")
                return None, "SENTOP Config sheet not found. Aborting."

        except Exception as e:   
            globalutils.show_stack_trace(str(e))
            return None, str(e)
 

        # ------------------------ GET SENTOP CONFIG DATA -------------------------

        data_sheet_name = None
        headers_row_index = None
        annotation = None
        user_stopwords = None
        
        # Config data starts in Row 4

        try:
            # Get datasheet name
            data_sheet_name = config_ws.cell(row=4, column=1).value
            if data_sheet_name:
                sentlog.info_keyval(f"Found data sheet name|{data_sheet_name}")
            else:
                return None, "No data sheet name found in SENTOP Config sheet."

            # Get headers row
            headers_row_index = config_ws.cell(row=4, column=2).value
            if headers_row_index:
                sentlog.info_keyval(f"Found headers row|{headers_row_index}")
            else:
                return None, "No headers row found in SENTOP Config sheet."

            # Get annotation
            annotation = config_ws.cell(row=4, column=3).value
            if annotation:
                sentlog.info_keyval(f"Found annotation|{annotation}")
            else:
                return None, "No annotation found in SENTOP Config sheet."

            # Get stop words
            user_stopwords = get_user_stopwords(config_ws, row=4, column=4)
            if user_stopwords:
                sentlog.info_keyval(f"Found user stop words|{user_stopwords}")
            else:
                sentlog.warn("No user stop words found in SENTOP Config sheet.")

        except Exception as e:   
            globalutils.show_stack_trace(str(e))
            return None, str(e)

        # ------------------------ GET DATA SHEET -------------------------

        data_ws = None
        try:
            data_ws = wb.get_sheet_by_name(data_sheet_name)
            if data_ws:
                sentlog.info_keyval(f"Found SENTOP data sheet|{data_sheet_name}")
            else:
                sentlog.error(f"SENTOP data sheet not found. Aborting.")
                return None, "SENTOP data sheet not found. Aborting."

        except Exception as e:   
            globalutils.show_stack_trace(str(e))
            return None, str(e)

        # ------------------------ GET DATA HEADERS -------------------------
        
        # ID header, if it exists, MUST use Excel Standard RED (#FF0000) font color.
        id_column_index = None
        id_header = None
        # Corpus header MUST use Excel Standard BLUE (#0070C0) font color.
        corpus_column_index = None
        corpus_header = None

        row = data_ws[headers_row_index]
        #col_num = 0
        headers = []
        for col_cell in row:
            print(f"Col {col_cell.column} header: {col_cell.value}")
            # Add column header
            headers.append(col_cell.value)
            #print(f"cell row: {col_cell.row}")
            #print(f"cell col: {col_cell.column}")
            if col_cell.font.color:
                #sentlog.info_p(f"Found font color")
                if 'rgb' in col_cell.font.color.__dict__:
                    font_color = col_cell.font.color.rgb
                    #sentlog.info_p(f"Found font color: {font_color}")
                    if font_color == 'FFFF0000':
                        id_header = col_cell.value
                        id_column_index = col_cell.column 
                        sentlog.info_keyval(f"Found ID column header|{id_header}")
                    elif font_color == 'FF0070C0':
                        corpus_header = col_cell.value
                        corpus_column_index = col_cell.column 
                        sentlog.info_keyval(f"Found corpus column header|{corpus_header}")
                        sentlog.info_keyval(f"Found corpus column index|{corpus_column_index}")
                    else:
                        sentlog.info_p(f"got non-red/blue column header: {font_color}")
            #else:
            #     sentlog.warn(f"No font color found for: {col_cell.value}")


        if not id_header:
            sentlog.warn(f"Could not find ID column header")

        if not corpus_header:
            sentlog.error(f"Could not find corpus column header")
            return None, "Could not find corpus column header."


        # --------------------- GET ALL STOP WORDS ---------------------

        all_stop_words = globalutils.get_all_stopwords(user_stopwords)

        # --------------------- GET ALL DATA CELLS ---------------------

        sentlog.info_h2("Data Validation")
        # Get XLSX data but ignore all rows that have invalid corpus data.       
        table_rows = []
        num_invalid_rows = 0
        for row in data_ws.iter_rows():

            row_num = row[0].row
            #print(f"Row num: {row_num}")
            if row_num <= headers_row_index:
                sentlog.warn(f"Skipping header row {row_num}")
                continue
            else:
                valid_corpus_data = False
                row_cols_data = []  # Row columns
                for col_cell in row:    
                    cell_value = col_cell.value
                    if col_cell.column == corpus_column_index:
                        # This is the row's corpus data
                        cleaned_text = text_validator.check(row_num, cell_value, all_stop_words) 
                        if cleaned_text:
                            valid_corpus_data = True
                            row_cols_data.append(cleaned_text)
                        else:
                            # Skip this row since text was invalid
                            # sentlog.warn(f"Skipping row {row_num} due to invalid corpus text.")
                            num_invalid_rows += 1
                    else:
                        # Add all other cells without cleaning
                        row_cols_data.append(cell_value)
                
                if valid_corpus_data:
                    #sentlog.info_p(f"Adding table row {row_num}")
                    table_rows.append(row_cols_data)
                    valid_corpus_data = False   # Reset valid corpus text to False

        sentlog.info_p(f"")
        sentlog.info_keyval(f"Number of invalid (ignored) rows|{num_invalid_rows}")
        sentlog.info_keyval(f"Number of valid rows|{len(table_rows)}")
        #globalutils.print_table(table_rows)

        # Get column slices
        row_id_list = None
        if id_column_index:
            row_id_list = globalutils.get_column(table_rows, id_column_index - 1)  #Python is 0-based
            print(f"row_id_list length: {len(row_id_list)}")
        #else:
        #    sentlog.warn(f"No ID column found. Using row numbers 0..{len(table_rows)} instead.")
        #    id_col = list(range(0, len(table_rows)))

        data_list = globalutils.get_column(table_rows, corpus_column_index - 1) #Python is 0-based
        print(f"data_list length: {len(data_list)}")

        if data_list:
            return XlsxData(kms_id, sentop_id, row_id_list, data_list, all_stop_words, annotation, table_rows, headers, headers_row_index), None
        else:
            return None, "Error getting data values from XLSX."

    except Exception as e:   
        globalutils.show_stack_trace(str(e))
        return None, str(e)

