import os
from contextlib import contextmanager,redirect_stderr,redirect_stdout
from os import devnull
import sys
import logging
import psutil
#from sklearn.feature_extraction import text
from topic_modeling import stopwords
from . import globalvars
import sentop_config as config
import nltk
from nltk.tokenize import word_tokenize
import string
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from . import sentop_log
import re

class Sentiments:
    def __init__(self, id, name, model_name, type, data_list):
        self.id = id
        self.name = name
        self.model_name = model_name
        self.type = type
        self.data_list = data_list

def get_column(matrix, i):
    #print(f"Slicing on column: {i}")
    columns = []
    for row in matrix:
        val = row[i]
        #print(f"Slice val: {val}")
        columns.append(val)
    #return [row[i] for row in matrix]
    return columns

def print_table(matrix):
    for row in matrix:
        print(">>>>>>>>>>>>>>>>>>>>")
        i = 1
        for col in row:
            print(f" {i}: {col}")
            i = i + 1


def show_stack_trace(error_msg):
    #print("Error: ", error_msg)
    exc_type, exc_obj, exc_tb = sys.exc_info()
    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
    #print(exc_type, fname, exc_tb.tb_lineno)
    sentlog = sentop_log.SentopLog()
    #sentlog.append(f"ERROR! {exc_type, fname, exc_tb.tb_lineno, error_msg}<br>")
    sentlog.error(f"{exc_type, fname, exc_tb.tb_lineno, error_msg}")

@contextmanager
def suppress_stdout_stderr():
    """A context manager that redirects stdout and stderr to devnull"""
    with open(devnull, 'w') as fnull:
        with redirect_stderr(fnull) as err, redirect_stdout(fnull) as out:
            yield (err, out)

# Disable logging
def block_logging():  
    # Disable logging  
    logging.config.dictConfig({
        'version': 1,
        'disable_existing_loggers': True,
    })


# Restore logging
def enable_logging():
    # Re-enable logging
    logging.config.dictConfig({
        'version': 1,
        'disable_existing_loggers': False,
    })


def get_memory():
    memory = psutil.Process(os.getpid()).memory_info().rss / 1024 ** 2
    mem_str = str(memory) + "MB"
    return mem_str


def get_memory():
    memory = psutil.Process(os.getpid()).memory_info().rss / 1024 ** 2
    mem_str = str(memory) + "MB"
    return mem_str


def get_stopwords_list(user_stop_words):
    frozen_stopwords = get_all_stopwords(user_stop_words)
    return list(frozen_stopwords)


# NOTE: user_stop_words are case sensitive. All other stopwords are case 
# insensitive.
def get_all_stopwords(user_stop_words):
    sentlog = sentop_log.SentopLog()

    lowercase_sentop_stopwords = [x.lower() for x in stopwords.stopwords_list]
    lowercase_user_stopwords = [y.lower() for y in user_stop_words]

    if user_stop_words:
        lowercase_sentop_stopwords.extend(lowercase_user_stopwords)

    #all_stop_words = text.ENGLISH_STOP_WORDS.union(lowercase_sentop_stopwords)
    #sentlog.info_p("trace 55")
    return lowercase_sentop_stopwords


def get_sentiment(id, sentiments):
    for r in sentiments:
        if r.id == id:
            #print(f"Found ID: {r.id}")
            return r


def generate_excel(id, annotation, row_id_list, data_list, sentiment_results, bertopic_results, lda_results, table_data, table_col_headers):
   
    sentlog = sentop_log.SentopLog()
    try:
        bert_sentence_topics = None
        bert_topics = None
        bert_duplicate_words = None

        if bertopic_results:
            bert_sentence_topics = bertopic_results.topic_per_row
            print(f"BERTopic rows: {len(bert_sentence_topics)}")
            bert_topics = bertopic_results.topics_list
            bert_duplicate_words = bertopic_results.duplicate_words_across_topics

        lda_sentence_topics = None
        lda_topics = None
        lda_duplicate_words = None

        if lda_results:
            lda_sentence_topics = lda_results.topic_per_row
            print(f"LDA rows: {len(lda_sentence_topics)}")
            lda_topics = lda_results.topics_list
            lda_duplicate_words = lda_results.duplicate_words_across_topics

        output_dir_path = config.data_dir_path.get("output")

        class3 = get_sentiment('class3', sentiment_results)
        class5 = get_sentiment('class5', sentiment_results)
        emotion1 = get_sentiment('emotion1', sentiment_results)
        offensive1 = get_sentiment('offensive1', sentiment_results) 
        emotion2 = get_sentiment('emotion2', sentiment_results)

        # Create results data
        rows = []
        for i in range(len(data_list)):
            row_data = []
            if (data_list[i]):
                #print(f"i = {i}")
                if row_id_list:
                    row_data.append(row_id_list[i])
                else:
                    row_data.append(i)

                row_data.append(data_list[i])
                if bert_sentence_topics:
                    row_data.append(bert_sentence_topics[i])
                else:
                    row_data.append("N/A")
                if lda_sentence_topics:
                    row_data.append(lda_sentence_topics[i])
                else:
                    row_data.append("N/A")
                row_data.append(class3.data_list[i])
                row_data.append(class5.data_list[i])
                row_data.append(emotion1.data_list[i])
                row_data.append(emotion2.data_list[i])
                row_data.append(offensive1.data_list[i])

                if table_data:
                    table_row = table_data[i]
                
                    #print(f"cols type: {type(table_row)}, cols length: {len(table_row)}, cols val: {table_row}")
                    vals = ""
                    for j in range (len(table_row)):
                        #print(f"j: {j}")
                        val = table_row[j]
                        if not val or val == 'None':
                            val = ""
                        
                        val = str(val)  # Make sure val is converted to string
                        val = val.strip('"')  # Remove double quotes from string
                        val = val.replace('"', "") # Remove single quotes from string
                        vals = vals + "\"" + str(val) + "\","  # Add double quotes around entire val
                        row_data.append(val)
                        #print(f"Vals: {vals}")

                rows.append(row_data)

        # Create results XLSX
        wb = Workbook()
        xlsx_out = output_dir_path + "\\" + id + "_results.xlsx"
        ws1 = wb.active
        ws1.title = "Results"

        # Create column headers
        headers = []
        if table_col_headers:
            # Get the table headers
            for header in table_col_headers:
                #print(f"Header start: {header}")
                if header is None:
                    header = "na"
                else:
                    header = header.replace("-","_")
                    header = header.replace(" ", "_")
                    header = re.sub("[^0-9a-zA-Z_]+", "", header)

        result_headers = ['ID', 'Document', 'BERTopic Topic', 'LDA Topic', class3.name, class5.name, emotion1.name, emotion2.name, offensive1.name]
        result_headers.extend(headers)
        ws1.append(result_headers)

        ws1['A1'].font = Font(bold=True)
        ws1['B1'].font = Font(bold=True)

        # Topic columns
        ws1['C1'].fill = PatternFill(start_color='FF66FF66', end_color='FF66FF66', fill_type='solid')
        ws1['C1'].font = Font(bold=True)
        ws1['D1'].fill = PatternFill(start_color='FF66FF66', end_color='FF66FF66', fill_type='solid')
        ws1['D1'].font = Font(bold=True)

        # Polarity sentiment columns
        ws1['E1'].fill = PatternFill(start_color='FF66FFFF', end_color='FF66FFFF', fill_type='solid')
        ws1['E1'].font = Font(bold=True)
        ws1['F1'].fill = PatternFill(start_color='FF66FFFF', end_color='FF66FFFF', fill_type='solid')
        ws1['F1'].font = Font(bold=True)

        # Emotion sentiment columns
        ws1['G1'].fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
        ws1['G1'].font = Font(bold=True)
        ws1['H1'].fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
        ws1['H1'].font = Font(bold=True)
        ws1['I1'].fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
        ws1['I1'].font = Font(bold=True)

        for i in range(len(rows)):
            ws1.append(rows[i])

        # Create Annotation XLSX sheet
        ws4 = wb.create_sheet(title="Annotation")
        fields = ['Annotation']
        annotation_list = []
        annotation_list.append(annotation)
        ws4.append(annotation_list)

        # Create BERTopic topics data
        if bert_topics:
            rows = []

            for i in range(len(bert_topics)):
                for j in range(len(bert_topics[i].words)):
                    row_data = []
                    row_data.append(bert_topics[i].topic_num)
                    row_data.append(bert_topics[i].words[j])
                    row_data.append(float(bert_topics[i].weights[j]))
                    rows.append(row_data)

            # Create BERTopic topics data XLSX sheet
            ws2 = wb.create_sheet(title="BERTopic")
            ws2.append(['Topic', 'Top Words', 'Weight'])
            for i in range(len(rows)):
                ws2.append(rows[i])

            # Create BERTopic non-overlapping topic words data
            rows = []
            for i in range(len(bert_topics)):
                for j in range(len(bert_topics[i].words)):
                    if not bert_topics[i].words[j] in bert_duplicate_words:
                        row_data = []
                        row_data.append(bert_topics[i].topic_num)
                        row_data.append(bert_topics[i].words[j])
                        row_data.append(float(bert_topics[i].weights[j]))
                        rows.append(row_data)

            # Create BERTopic non-overlapping topics data XLSX sheet
            ws2 = wb.create_sheet(title="BERTopic Non-Overlapping Topics")
            ws2.append(['Topic', 'Top Words', 'Weight'])
            for i in range(len(rows)):
                ws2.append(rows[i])  

        # Create LDA topics data
        if lda_topics:
            rows = []
            for i in range(len(lda_topics)):
                #print("LDA i: ", i)
                for j in range(len(lda_topics[i].words)):
                    #print("LDA j: ", j)
                    row_data = []
                    row_data.append(lda_topics[i].topic_num)
                    row_data.append(lda_topics[i].words[j])
                    row_data.append(float(lda_topics[i].weights[j]))
                    rows.append(row_data)

            # Create LDA topics data XLSX sheet
            ws3 = wb.create_sheet(title="LDA")
            fields = ['Topic', 'Top Words', 'Weight']
            ws3.append(fields)
            for i in range(len(rows)):
                ws3.append(rows[i])

            # Create LDA non-overlapping topics words data
            rows = []
            for i in range(len(lda_topics)):
                #print("LDA i: ", i)
                for j in range(len(lda_topics[i].words)):
                    if not lda_topics[i].words[j] in lda_duplicate_words:
                        row_data = []
                        row_data.append(lda_topics[i].topic_num)
                        row_data.append(lda_topics[i].words[j])
                        row_data.append(float(lda_topics[i].weights[j]))
                        rows.append(row_data)

            # Create LDA topics data XLSX sheet
            ws3 = wb.create_sheet(title="LDA Non-Overlapping Topics")
            fields = ['Topic', 'Top Words', 'Weight']
            ws3.append(fields)
            for i in range(len(rows)):
                ws3.append(rows[i])

        # Save XLSX
        wb.save(filename = xlsx_out)
    except Exception as e:
        show_stack_trace(e)

