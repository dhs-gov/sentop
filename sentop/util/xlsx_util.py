"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import logging
from . import log_util
import string


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging
import traceback
from datetime import datetime

from ..sentiment_analysis import class3 as sent_class3
from ..sentiment_analysis import class5 as sent_class5
from ..sentiment_analysis import emotion2 as sent_emotion2
from ..sentiment_analysis import offensive1 as sent_offensive1


class XlsxDataIn ():

    def __init__(self, table_data, headers_row_index, headers, \
        id_column_letter, id_column_index, narrative_column_letter, narrative_column_index, annotation, user_stopwords):

        self.table_data = table_data
        self.headers_row_index = headers_row_index
        self.headers = headers
        self.id_column_letter = id_column_letter
        self.id_column_index = id_column_index
        self.narrative_column_letter = narrative_column_letter
        self.narrative_column_index = narrative_column_index
        self.annotation = annotation
        self.user_stopwords = user_stopwords


    def show_info(self):
        logger = logging.getLogger()

        if self.table_data:
            logger.debug(f"Table data rows: {len(self.table_data)}, cols: {len(self.headers)}")
        else:
            logger.error("No table data for this object.")

        if self.annotation:
            logger.debug(f"Annotation: {self.annotation}")
        else:
            logger.error("No annotation for this data.")

        if self.user_stopwords:
            logger.debug(f"User stop words: {len(self.user_stopwords)}")
        else:
            logger.error("No user stop words for this data.")


def get_user_stopwords(config_ws, row, column):
    logger = logging.getLogger()
    user_stop_words = []
    try:
        done = False
        while not done:
            stop_word = config_ws.cell(row, column).value
            if stop_word:
                logger.debug(f"Found stop word: {stop_word}")
                user_stop_words.append(stop_word)
                row += 1 
            else:
                done = True

        return user_stop_words

    except Exception as e:
        print(traceback.format_exc())
        logger.error("Error getting user user stop words.")
        return None


def column_letter_to_number(c):
    """Return number corresponding to excel-style column."""
    number=-25
    for l in c:
        if not l in string.ascii_letters:
            return False
        number+=ord(l.upper())-64+25
    return number


def get_data(input_file):
    # Get data from XLSX file.
    logger = logging.getLogger()

    try:
        # ------------------------ GET WORKBOOK -------------------------

        wb = load_workbook(filename=input_file, data_only=True)

        # ------------------------ GET CONFIG SHEET -------------------------

        # XLSX files MUST have a SenTop config sheet.
        config_ws = None
        try:
            config_ws = wb.get_sheet_by_name('SENTOP Config')
            if not config_ws:
                msg = 'SENTOP Config sheet not found. Aborting.'
                logger.error(f'{msg}')
                return None, msg
        except Exception as e:   
            print(traceback.format_exc())
            return None, str(e)
 
        # Get SenTop config data.
        data_sheet_name = None
        headers_row_index = None
        id_column_letter = None
        id_column_index = None
        narrative_column_letter = None
        narrative_column_index = None
        annotation = None
        user_stopwords = None
        
        # Config data starts in Row 4
        try:
            logger.info("Getting config data")
            # Get mandatory data sheet name
            data_sheet_name = config_ws.cell(row=4, column=1).value
            if data_sheet_name:
                logger.info(f"Found data sheet name: {data_sheet_name}")
            else:
                return None, "No data sheet name found in SENTOP Config sheet."

            # Get mandatory headers row
            headers_row_index = config_ws.cell(row=4, column=2).value
            if headers_row_index:
                logger.info(f"Found headers row: {headers_row_index}")
            else:
                return None, "No headers row found in SENTOP Config sheet."

             # Get optional ID column
            id_column_letter = config_ws.cell(row=4, column=3).value
            if id_column_letter:
                logger.info(f"Found ID column: {id_column_letter}")
                id_column_index = column_letter_to_number(id_column_letter)
                logger.info(f"Found ID column index: {id_column_index}")
            else:
                logger.info(f"No optional ID column found.")

            # Get mandatory narrative column
            narrative_column_letter = config_ws.cell(row=4, column=4).value
            if narrative_column_letter:
                logger.info(f"Found narratives column: {narrative_column_letter}")
                narrative_column_index = column_letter_to_number(narrative_column_letter)
                logger.info(f"Found narratives index: {narrative_column_index}")
            else:
                return None, "No narratives column found in SENTOP Config sheet."               

            # Get optional annotation
            annotation = config_ws.cell(row=4, column=5).value
            if annotation:
                logger.info(f"Found annotation: {annotation}")
            else:
                logger.info(f"No annotation: {annotation}")

            # Get stop words
            user_stopwords = get_user_stopwords(config_ws, row=4, column=6)
            if user_stopwords:
                logger.info(f"Found user stop words: {user_stopwords}")
            else:
                logger.info("No user stop words found in SENTOP Config sheet.")

        except Exception as e:   
            print(traceback.format_exc())
            return None, str(e)

        # ------------------------ GET DATA SHEET -------------------------
        logger.info("Getting data sheet")
        data_ws = None
        try:
            data_ws = wb.get_sheet_by_name(data_sheet_name)
            if data_ws:
                logger.info(f"Found SENTOP data sheet: {data_sheet_name}")
            else:
                logger.error(f"SENTOP data sheet not found. Aborting.")
                return None, "SENTOP data sheet not found. Aborting."

        except Exception as e:   
            print(traceback.format_exc())
            return None, str(e)

        # ------------------------ GET TABLE HEADERS -------------------------
        
        id_column_index = None
        id_header = None
        corpus_column_index = None
        corpus_header = None

        row = data_ws[headers_row_index]
        headers = []
        for col_cell in row:
            col_letter = get_column_letter(col_cell.column)
            headers.append(col_cell.value)
            if id_column_letter:
                if id_column_letter == col_letter:
                        id_header = col_cell.value
                        id_column_index = col_cell.column 
                        logger.debug(f"Found ID column header: {id_header}")
            if narrative_column_letter:
                if narrative_column_letter == col_letter:
                        corpus_header = col_cell.value
                        corpus_column_index = col_cell.column 
                        logger.debug(f"Found corpus column header: {corpus_header}")
                        logger.debug(f"Found corpus column index: {corpus_column_index}")

        if not id_header:
            logger.warning(f"Could not find ID column header")

        if not corpus_header:
            logger.error(f"Could not find corpus column header")
            return None, "Could not find corpus column header."


        # --------------------- GET ALL STOP WORDS ---------------------

        #Stopwords are acquired in _sentop.py


        # --------------------- GET ALL DATA CELLS ---------------------

        # Get XLSX data.       
        table_data = []
        num_invalid_rows = 0
        for row in data_ws.iter_rows():
            row_num = row[0].row
            #print(f"Row num: {row_num}")
            if row_num <= headers_row_index:
                logger.debug(f"Skipping header row {row_num}")
                continue
            else:
                row_cols_data = []  # Row columns
                for col_cell in row:    
                    cell_value = col_cell.value
                    if cell_value:
                        row_cols_data.append(cell_value)
                    else:
                        row_cols_data.append('')
                # Add row columns to table rows
                table_data.append(row_cols_data)
        
        return XlsxDataIn(table_data, headers_row_index, headers, id_column_letter, id_column_index, narrative_column_letter, narrative_column_index, annotation, user_stopwords), None

    except Exception as e:   
        print(traceback.format_exc())
        return None, str(e)



def generate_excel(results_name, preprocessing_results, annotation, 
    row_id_list, docs, sentiments, lda_results, bertopic_results, analyses_results, 
    RESULTS_DIR):
    logger = logging.getLogger()

    try:
        bert_sentence_topics = None
        bert_topics = None
        bert_duplicate_words = None

        if bertopic_results:
            bert_sentence_topics = bertopic_results.topic_per_row
            logger.info(f"BERTopic rows: {len(bert_sentence_topics)}")
            bert_topics = bertopic_results.topics_list
            bert_duplicate_words = bertopic_results.duplicate_words_across_topics

        lda_sentence_topics = None
        lda_topics = None
        lda_duplicate_words = None

        if lda_results:
            lda_sentence_topics = lda_results.topic_per_row
            print(f"LDA rows: {len(lda_sentence_topics)}")
            lda_topics = lda_results.topics_list
            lda_duplicate_words = lda_results.duplicate_words_across_topics

        output_dir_path = RESULTS_DIR

        class3 = sentiments.class3
        class5 = sentiments.class5
        emotion1 = sentiments.emotion1
        emotion2 = sentiments.emotion2
        offensive1 = sentiments.offensive1

        # Create results data
        rows = []
        for i in range(len(docs)):
            row_data = []
            if (docs[i]):
                # Write row ID
                if row_id_list:
                    row_data.append(row_id_list[i])
                else:
                    row_data.append(i)

                # Write preprocessed document
                row_data.append(docs[i])

                # Write preprocessor result
                row_data.append(preprocessing_results[i])

                # Write BERTopic topic
                if bert_sentence_topics:
                    row_data.append(bert_sentence_topics[i])
                else:
                    row_data.append("N/A")

                # Write LDA topic
                if lda_sentence_topics:
                    row_data.append(lda_sentence_topics[i])
                else:
                    row_data.append("N/A")

                # Write class3
                if class3:
                    row_data.append(class3[i])
                else:
                    row_data.append("N/A")    

                # Write class 5
                if class5:
                    row_data.append(class5[i])
                else:
                    row_data.append("N/A") 

                # Write emotion1
                if emotion1:
                    row_data.append(emotion1[i])
                else:
                    row_data.append("N/A")  

                # Write emotion2
                if emotion2:
                    row_data.append(emotion2[i])
                else:
                    row_data.append("N/A") 

                # Write offensive1
                if offensive1:
                    row_data.append(offensive1[i])
                else:
                    row_data.append("N/A") 


                rows.append(row_data)

        # Create results XLSX
        wb = Workbook()
        #job_id = datetime.now().strftime('%m%d%Y_%H%M%S')
        xlsx_out = RESULTS_DIR + "\\sentop_results_" + str(results_name) + ".xlsx"
        ws1 = wb.active
        ws1.title = "Results"

        # Create column headers
        result_headers = ['ID', 'Document', 'Preproc Status', 'BERTopic Topic', 'LDA Topic', 'Class-3', 'Class-5', 'Emotion-1', 'Emotion-2', 'Offensive-1']
        #result_headers.extend(headers)
        ws1.append(result_headers)

        ws1['A1'].font = Font(bold=True)
        ws1['B1'].font = Font(bold=True)
        ws1['C1'].font = Font(bold=True)

        # Topic columns
        ws1['D1'].fill = PatternFill(start_color='FF66FF66', end_color='FF66FF66', fill_type='solid')
        ws1['D1'].font = Font(bold=True)
        ws1['E1'].fill = PatternFill(start_color='FF66FF66', end_color='FF66FF66', fill_type='solid')
        ws1['E1'].font = Font(bold=True)

        # Polarity sentiment columns
        ws1['F1'].fill = PatternFill(start_color='FF66FFFF', end_color='FF66FFFF', fill_type='solid')
        ws1['F1'].font = Font(bold=True)
        ws1['G1'].fill = PatternFill(start_color='FF66FFFF', end_color='FF66FFFF', fill_type='solid')
        ws1['G1'].font = Font(bold=True)

        # Emotion sentiment columns
        ws1['H1'].fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
        ws1['H1'].font = Font(bold=True)
        ws1['I1'].fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
        ws1['I1'].font = Font(bold=True)
        ws1['J1'].fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
        ws1['J1'].font = Font(bold=True)

        for i in range(len(rows)):
            ws1.append(rows[i])

        # Create Annotation XLSX sheet
        ws4 = wb.create_sheet(title="Annotation")
        fields = ['Annotation']
        annotation_list = []
        annotation_list.append(annotation)
        ws4.append(annotation_list)

        # Create BERTopic topics data
        bert_rows = []
        if bert_topics:

            for i in range(len(bert_topics)):
                for j in range(len(bert_topics[i].words)):
                    row_data = []
                    row_data.append(bert_topics[i].topic_num)
                    row_data.append(bert_topics[i].words[j])
                    row_data.append(float(bert_topics[i].weights[j]))
                    bert_rows.append(row_data)

            # Create BERTopic topics data XLSX sheet
            ws2 = wb.create_sheet(title="BERTopic")
            ws2.append(['Topic', 'Top Words', 'Weight'])
            for i in range(len(bert_rows)):
                ws2.append(bert_rows[i])

            # Create BERTopic non-overlapping topic words data
            bert_rows_nonoverlapping = []
            for i in range(len(bert_topics)):
                for j in range(len(bert_topics[i].words)):
                    if not bert_topics[i].words[j] in bert_duplicate_words:
                        row_data = []
                        row_data.append(bert_topics[i].topic_num)
                        row_data.append(bert_topics[i].words[j])
                        row_data.append(float(bert_topics[i].weights[j]))
                        bert_rows_nonoverlapping.append(row_data)

            # Create BERTopic non-overlapping topics data XLSX sheet
            ws2 = wb.create_sheet(title="BERTopic Non-Overlapping Topics")
            ws2.append(['Topic', 'Top Words', 'Weight'])
            for i in range(len(bert_rows_nonoverlapping)):
                ws2.append(bert_rows_nonoverlapping[i])  

        # Create LDA topics data
        lda_rows = []
        if lda_topics:
            for i in range(len(lda_topics)):
                #print("LDA i: ", i)
                for j in range(len(lda_topics[i].words)):
                    #print("LDA j: ", j)
                    row_data = []
                    row_data.append(lda_topics[i].topic_num)
                    row_data.append(lda_topics[i].words[j])
                    row_data.append(float(lda_topics[i].weights[j]))
                    lda_rows.append(row_data)

            # Create LDA topics data XLSX sheet
            ws3 = wb.create_sheet(title="LDA")
            fields = ['Topic', 'Top Words', 'Weight']
            ws3.append(fields)
            for i in range(len(lda_rows)):
                ws3.append(lda_rows[i])

            # Create LDA non-overlapping topics words data
            lda_rows_nonoverlapping = []
            for i in range(len(lda_topics)):
                #print("LDA i: ", i)
                for j in range(len(lda_topics[i].words)):
                    if not lda_topics[i].words[j] in lda_duplicate_words:
                        row_data = []
                        row_data.append(lda_topics[i].topic_num)
                        row_data.append(lda_topics[i].words[j])
                        row_data.append(float(lda_topics[i].weights[j]))
                        lda_rows_nonoverlapping.append(row_data)

            # Create LDA topics data XLSX sheet
            ws3 = wb.create_sheet(title="LDA Non-Overlapping Topics")
            fields = ['Topic', 'Top Words', 'Weight']
            ws3.append(fields)
            for i in range(len(lda_rows_nonoverlapping)):
                ws3.append(lda_rows_nonoverlapping[i])


        #------------------------------ ANALYSES ---------------------------------


        def write_sentiments(title, item_counts_list, sent_type):
            print(f"Writing {title}")
            rows = []
            for i, sentiment_count in enumerate(item_counts_list):
                sentiment_label = sent_type.get_sentiment_label(i)
                row_data = []
                row_data.append(sentiment_label)
                row_data.append(sentiment_count)
                rows.append(row_data)
            # Create sheet
            ws2 = wb.create_sheet(title=title)
            # Create header row
            ws2.append(['Sentiment', 'Count'])
            # Bold header row 1
            for cell in ws2["1:1"]:
                cell.font = Font(bold=True)
            for i in range(len(rows)):
                ws2.append(rows[i])
        

        def write_topics(title, occurance_list, topic_numbers, topics):
            print(f"Writing {title}")
            rows = []
            for i, topic_count in enumerate(occurance_list):
                topic_number = topic_numbers[i]
                topic = topics[i]
                topic_label = ', '.join(topic.words)
                if (topic_number < 10):
                    # This is to make sure charts labels are ordered properly in Excel
                    topic_label = " " + str(topic_number) + ": " + topic_label
                else:
                    topic_label = str(topic_number) + ": " + topic_label
                row_data = []
                row_data.append(topic_number)
                row_data.append(topic_label)
                row_data.append(topic_count)
                rows.append(row_data)
            # Create offensive-1 sheet
            ws2 = wb.create_sheet(title=title)
            # Create header row
            ws2.append(['Topic Num', 'Topic', 'Count'])
            # Bold header row 1
            for cell in ws2["1:1"]:
                cell.font = Font(bold=True)
            # Write data
            for i in range(len(rows)):
                ws2.append(rows[i])


        def write_topics_with_sentiments(title, occurance_list, topics, sent_type):
            print(f"Writing {title}")
            rows = []
            for i, topic_sent in enumerate(occurance_list):
                #topic_number = analyses_results.lda_topics_list[i]
                topic = topics[i]
                topic_label = ', '.join(topic.words)
                topic_label = str(topic_sent[0]) + ": " + topic_label
                row_data = []
                row_data.append(topic_sent[0])
                row_data.append(topic_label)
                for i in range(1, len(topic_sent)):
                    row_data.append(topic_sent[i])
  
                rows.append(row_data)
            # Create sheet
            ws2 = wb.create_sheet(title=title)
            # Create headers
            header_list1 = ['Topic Num', 'Topic']
            header_list2 = list(sent_type.mappings.values())
            headers = header_list1 + header_list2
            ws2.append(headers)
            # Bold header row 1
            for cell in ws2["1:1"]:
                cell.font = Font(bold=True)
            # Write the data
            for i in range(len(rows)):
                ws2.append(rows[i])


        def write_sentiments_with_topics(title, occurance_list, topics):
            print(f"Writing {title}")
            rows = []
            for i, sent_topic in enumerate(occurance_list):
                row_data = []
                row_data.append(sent_topic[0]) # Sentiment label
                for i in range(1, len(sent_topic)):
                    row_data.append(sent_topic[i]) # Topic numbers
  
                rows.append(row_data)
            # Create sheet
            ws2 = wb.create_sheet(title=title)
            # Create headers
            header_list1 = ['Sentiment']
            header_list2 = topics
            headers = header_list1 + header_list2
            ws2.append(headers)
            # Bold header row 1
            for cell in ws2["1:1"]:
                cell.font = Font(bold=True)
            # Write the data
            for i in range(len(rows)):
                ws2.append(rows[i])


        # -------------------- Sentiments

        # Create class 3
        if analyses_results.class3_counts:
            write_sentiments("Analyses_Class-3", analyses_results.class3_counts, sent_class3)

        # Create class 5
        if analyses_results.class5_counts:
            write_sentiments("Analyses_Class-5", analyses_results.class5_counts, sent_class5)

        # Create Emotion-2
        if analyses_results.emotion2_counts:
            write_sentiments("Analyses_Emotion-2", analyses_results.emotion2_counts, sent_emotion2)

        # Create Offensive-1
        if analyses_results.offensive1_counts:
            write_sentiments("Analyses_Offensive-1", analyses_results.offensive1_counts, sent_offensive1)

        # -------------------- Topics

        # LDA
        if analyses_results.lda_occurrence_list:
            write_topics("Analyses_LDA", \
                analyses_results.lda_occurrence_list, \
                analyses_results.lda_topics_list, lda_topics)

        # BERTopic
        if analyses_results.bertopic_occurrence_list:
            write_topics("Analyses_BERTopic", 
                analyses_results.bertopic_occurrence_list, \
                analyses_results.bertopic_topics_list, bert_topics)

        # -------------------- LDA Sentiments

        # LDA 3-class
        if analyses_results.lda_3class:
            write_topics_with_sentiments("Analyses_LDA-3Class", \
                analyses_results.lda_3class, \
                lda_topics, sent_class3)


        # LDA 5-class
        if analyses_results.lda_5class:

            write_topics_with_sentiments("Analyses_LDA-5Class", \
                analyses_results.lda_5class, \
                lda_topics, sent_class5)


        # LDA Emotion-2
        if analyses_results.lda_emotion2:

            write_topics_with_sentiments("Analyses_LDA-Emotion2", \
                analyses_results.lda_emotion2, \
                lda_topics, sent_emotion2)


        # LDA Offensive-1
        if analyses_results.lda_offensive1:

            write_topics_with_sentiments("Analyses_LDA-Offensive1", \
                analyses_results.lda_offensive1, \
                lda_topics, sent_offensive1)


        # -------------------- Sentiment LDA

        # 3-class LDA Topics
        if analyses_results.class3_lda:
            write_sentiments_with_topics("Analyses_Class3-LDA", \
                analyses_results.class3_lda, \
                analyses_results.lda_topics_list)


        # 5-class LDA Topics
        if analyses_results.class5_lda:
            write_sentiments_with_topics("Analyses_Class5-LDA", \
                analyses_results.class5_lda, \
                analyses_results.lda_topics_list)


        # Emotion-2 LDA Topics
        if analyses_results.emotion2_lda:

            write_sentiments_with_topics("Analyses_Emotion2-LDA", \
                analyses_results.emotion2_lda, \
                analyses_results.lda_topics_list)

        # Offensive-1 LDA Topics
        if analyses_results.offensive1_lda:

            write_sentiments_with_topics("Analyses_Offensive1-LDA", \
                analyses_results.offensive1_lda, \
                analyses_results.lda_topics_list)


       # -------------------- BERTopic Sentiments

        # BERTopic 3-class
        if analyses_results.bertopic_3class:

            write_topics_with_sentiments("Analyses_BERTopic-3Class", \
                analyses_results.bertopic_3class, \
                bert_topics, sent_class3)


        # BERTopic 5-class
        if analyses_results.bertopic_5class:

            write_topics_with_sentiments("Analyses_BERTopic-5Class", \
                analyses_results.bertopic_5class, \
                bert_topics, sent_class5)
                

        # BERTopic Emotion-2
        if analyses_results.bertopic_emotion2:

            write_topics_with_sentiments("Analyses_BERTopic-Emotion2", \
                analyses_results.bertopic_emotion2, \
                bert_topics, sent_emotion2)


        # BERTopic Offensive-1
        if analyses_results.bertopic_offensive1:

            write_topics_with_sentiments("Analyses_BERTopic-Offensive1", \
                analyses_results.bertopic_offensive1, \
                bert_topics, sent_offensive1)
                

       # -------------------- Sentiment BERTopic

        # 3-class BERTopic Topics
        if analyses_results.class3_bertopic:

           write_sentiments_with_topics("Analyses_Class3-BERTopic", \
                analyses_results.class3_bertopic, \
                analyses_results.bertopic_topics_list)


        # 5-class BERTopic Topics
        if analyses_results.class5_bertopic:

           write_sentiments_with_topics("Analyses_Class5-BERTopic", \
                analyses_results.class5_bertopic, \
                analyses_results.bertopic_topics_list)


        # Emotion-2 BERTopic Topics
        if analyses_results.emotion2_bertopic:

           write_sentiments_with_topics("Analyses_Emotion2-BERTopic", \
                analyses_results.emotion2_bertopic, \
                analyses_results.bertopic_topics_list)


        # Offensive-1 LDA Topics
        if analyses_results.offensive1_bertopic:

           write_sentiments_with_topics("Analyses_Offensive1-BERTopic", \
                analyses_results.offensive1_bertopic, \
                analyses_results.bertopic_topics_list)


        # Save XLSX
        wb.save(filename = xlsx_out)

        logger.info(f"Wrote Excel results to: {xlsx_out}")
    except Exception as e:
        print(traceback.format_exc())
        return None, str(e)
"""
